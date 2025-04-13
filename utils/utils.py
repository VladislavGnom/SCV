import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.contrib.auth.models import Group
from django.core.exceptions import ObjectDoesNotExist

def extract_filename_substring(input_string):
    # Шаблон для имени файла: допустим, это имя файла с расширением
    pattern = r'\b\w+\.\w+\b'
    
    # Поиск совпадений в строке
    match = re.search(pattern, input_string)
    
    if match:
        return match.group(0)  # Возвращаем найденное имя файла
    else:
        return None  # Если совпадений нет, возвращаем None

# # Пример использования
# input_str = "Путь к файлу: /папка/файл_пример.txt"
# result = extract_filename_substring(input_str)
# print(result)  # Выведет: файл_пример.txt


from bs4 import BeautifulSoup

def clean_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    return str(soup)

# raw_html = "<div><p>Some text"
# cleaned_html = clean_html(raw_html)


# CREATE EXCEL FILE

import os
import pandas as pd


def create_excel_table(data: dict[str, list], filename: str, sheet_name: str = 'Sheet1'):
    df = pd.DataFrame(data)

    if not os.path.exists('static/excel_files'):
        os.makedirs('static/excel_files')

    filepath = os.path.join('static/excel_files', filename)
    df.to_excel(filepath, index=False, sheet_name=sheet_name)

    return filepath


def restyles_excel_file(filename: str, sheet_name: str = 'Sheet1'):
    filepath = os.path.join('static/excel_files', filename)
     # Открываем существующий Excel файл для редактирования
    wb = load_workbook(filepath)
    sheet = wb.active

    # Выравнивание текста по центру
    alignment = Alignment(horizontal='center', vertical='center')
    font_style = Font(bold=True, color='FFFFFF')  # Жирный белый текст
    fill_color = PatternFill(start_color='4F5249', end_color='4F5249', fill_type='solid')  # Серый фон

    # Определение стилей обводки
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))


    # Применение выравнивания к ячейкам
    # Применение стилей к столбцу A
    for cell in sheet['A']: 
        cell.alignment = alignment
        cell.font = font_style
        cell.fill = fill_color
        cell.border = thin_border

    # Применение стилей к столбцу B
    for cell in sheet['B']: 
        cell.alignment = alignment
        cell.font = font_style
        cell.fill = fill_color
        cell.border = thin_border

    
    # Применение стилей к столбцу C
    for cell in sheet['C']: 
        cell.alignment = alignment
        cell.font = font_style
        cell.fill = fill_color
        cell.border = thin_border

    # Применение стилей к столбцу D
    for cell in sheet['D']: 
        cell.alignment = alignment
        cell.font = font_style
        cell.fill = fill_color
        cell.border = thin_border

    # Применение стилей к столбцу E
    for cell in sheet['E']: 
        cell.alignment = alignment
        cell.font = font_style
        cell.fill = fill_color
        cell.border = thin_border
    
    fill_color = PatternFill(start_color='71a102', end_color='71a102', fill_type='solid')  # Лаймовый фон


    # change fill material of title
    sheet["A1"].fill = fill_color
    sheet["B1"].fill = fill_color
    sheet["C1"].fill = fill_color
    sheet["D1"].fill = fill_color
    sheet["E1"].fill = fill_color

    # Изменяем ширину столбцов
    sheet.column_dimensions['A'].width = 25  # Ширина для колонки "Ученик"
    sheet.column_dimensions['B'].width = 35  # Ширина для колонки "Название теста"
    sheet.column_dimensions['C'].width = 30  # Ширина для колонки "ID заданий"
    sheet.column_dimensions['D'].width = 40  # Ширина для колонки "Кол-во правильных ответов"
    sheet.column_dimensions['E'].width = 40  # Ширина для колонки "Процент выполнения работы"

    # Сохраняем изменения
    wb.save(filepath)

def get_group_by_name(name_of_group: str) -> Group | None:
    try:
        group = Group.objects.get(name=name_of_group)
        return group 
    except ObjectDoesNotExist as exc:
        return None

# -------------------- USER APP ---------------
# help function 
# отбирает все группы в которые входит пользователь, который был передан как аргумент
def get_user_groups(user):
    return [group.id for group in user.groups.all()]


def str_to_int(obj):
    return int(obj)